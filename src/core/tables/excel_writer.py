import openpyxl
from openpyxl.utils import get_column_letter #good lord it's not my firts time with openpexl ong
from typing import List, Dict
from src.core.tables.table import Table
from src.core.cells.reference_handler import ReferenceHandler

class ExcelWriter:
    def __init__(self, output_file: str):
        self.output_file = output_file
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)
        self.table_id_map: Dict[str, str] = {}

    def write_tables(self, tables: List[Table]):
        # first for loop: creating sheets and map IDs to names secodn loop: actually writing stuff
        for table in tables:
            sheet_name = self._sanitize_sheet_name(table.name)
            original_name = sheet_name
            counter = 1
            while sheet_name in self.workbook.sheetnames:
                sheet_name = f"{original_name[:28]}_{counter}"
                counter += 1

            self.workbook.create_sheet(title=sheet_name)
            self.table_id_map[table.id] = sheet_name

        for table in tables:
            sheet_name = self.table_id_map[table.id]
            ws = self.workbook[sheet_name]

            for cell in table.cells:
                value = cell.get_value()

                if isinstance(cell, ReferenceHandler) and isinstance(value, str) and value.startswith("LINK:"):
                    #soo the format is LINK:table_id!cell_address
                    try:
                        _, rest = value.split(":", 1)
                        ref_table_id, ref_address = rest.split("!", 1)
                        if ref_table_id in self.table_id_map:
                            ref_sheet_name = self.table_id_map[ref_table_id]
                            value = f"='{ref_sheet_name}'!{ref_address}"
                        else:
                            value = f"#REF!{ref_table_id}"
                    except ValueError:
                         value = "#ERROR!"

                #cell.coordinate is like "A1"
                ws[cell.coordinate] = value

        self.workbook.save(self.output_file)

    def _sanitize_sheet_name(self, name: str) -> str:
        #had to google, apparently Excel sheet names cannot contain : \ / ? * [ ] aaand max 31 chars
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '')
        return name[:31]