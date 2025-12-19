import unittest
from unittest.mock import MagicMock, patch
import os
import sys
from datetime import datetime

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.api.client import APIClient
from src.api.models import ApiTable, ApiCell, CellType, CellReference, FormatType
from src.core.cells.value_cell import ValueCell
from src.core.cells.formula_cell import FormulaCell
from src.core.cells.reference_handler import ReferenceHandler
from src.core.tables.table import Table
from src.core.tables.excel_writer import ExcelWriter

class TestModels(unittest.TestCase):
    def test_api_cell_creation(self):
        cell = ApiCell(id="c1", row=1, column=1, cell_type=CellType.VALUE, value=100)
        self.assertEqual(cell.value, 100)
        self.assertEqual(cell.excel_address, "A1")

    def test_api_cell_column_conversion(self):
        cell = ApiCell(id="c1", row=1, column=26, cell_type=CellType.VALUE)
        self.assertEqual(cell.excel_address, "Z1")
        cell.column = 27
        self.assertEqual(cell.excel_address, "AA1")

    def test_api_table_creation(self):
        cell = ApiCell(id="c1", row=1, column=1, cell_type=CellType.VALUE, value=100)
        table = ApiTable(id="t1", name="Test Table", cells=[cell])
        self.assertEqual(table.name, "Test Table")
        self.assertEqual(len(table.cells), 1)

class TestCells(unittest.TestCase):
    def test_value_cell(self):
        api_cell = ApiCell(id="c1", row=1, column=1, cell_type=CellType.VALUE, value="test")
        cell = ValueCell(api_cell)
        self.assertEqual(cell.get_value(), "test")
        self.assertEqual(cell.coordinate, "A1")

    def test_formula_cell(self):
        api_cell = ApiCell(id="c1", row=1, column=1, cell_type=CellType.FORMULA, formula="SUM(A1:A2)")
        cell = FormulaCell(api_cell)
        self.assertEqual(cell.get_value(), "=SUM(A1:A2)")

        api_cell_eq = ApiCell(id="c1", row=1, column=1, cell_type=CellType.FORMULA, formula="=SUM(A1:A2)")
        cell_eq = FormulaCell(api_cell_eq)
        self.assertEqual(cell_eq.get_value(), "=SUM(A1:A2)")

    def test_reference_handler(self):
        ref = CellReference(table_id="t2", cell_address="B2")
        api_cell = ApiCell(id="c1", row=1, column=1, cell_type=CellType.LINK, reference=[ref])
        cell = ReferenceHandler(api_cell)
        self.assertEqual(cell.get_value(), "LINK:t2!B2")
        self.assertEqual(cell.get_referenced_table_ids(), ["t2"])

class TestTable(unittest.TestCase):
    def test_table_processing(self):
        cells = [
            ApiCell(id="c1", row=1, column=1, cell_type=CellType.VALUE, value="v"),
            ApiCell(id="c2", row=1, column=2, cell_type=CellType.FORMULA, formula="=A1"),
            ApiCell(id="c3", row=2, column=1, cell_type=CellType.LINK, reference=[CellReference(table_id="t2", cell_address="A1")]),
            ApiCell(id="c4", row=3, column=1, cell_type=CellType.EMPTY)
        ]
        api_table = ApiTable(id="t1", name="Test", cells=cells)
        table = Table(api_table)

        self.assertEqual(len(table.cells), 3) # Empty cell is skipped
        self.assertIsInstance(table.cells[0], ValueCell)
        self.assertIsInstance(table.cells[1], FormulaCell)
        self.assertIsInstance(table.cells[2], ReferenceHandler)
        self.assertEqual(table.get_linked_table_ids(), ["t2"])

class TestAPIClient(unittest.TestCase):
    def setUp(self):
        self.client = APIClient(token="test_token")

    @patch('src.api.client.requests.get')
    def test_get_tables(self, mock_get):
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.json.return_value = {"tables": [{"id": "t1", "name": "Table 1"}]}
        mock_get.return_value = mock_response

        tables = self.client.get_tables()
        self.assertEqual(len(tables), 1)
        self.assertEqual(tables[0]['id'], "t1")

    @patch('src.api.client.requests.get')
    def test_get_table(self, mock_get):
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.json.return_value = {
            "id": "t1",
            "name": "Table 1",
            "cells": [{"id": "c1", "row": 1, "column": 1, "cell_type": "value", "value": 10}]
        }
        mock_get.return_value = mock_response

        table = self.client.get_table("t1")
        self.assertIsInstance(table, ApiTable)
        self.assertEqual(table.id, "t1")
        self.assertEqual(len(table.cells), 1)

    def test_mock_mode(self):
        mock_data = {"t1": ApiTable(id="t1", name="Mock", cells=[])}
        self.client.set_mock_mode(mock_data)

        tables = self.client.get_tables()
        self.assertEqual(len(tables), 1)
        self.assertEqual(tables[0]['id'], "t1")

        table = self.client.get_table("t1")
        self.assertEqual(table.name, "Mock")

class TestExcelWriter(unittest.TestCase):
    def test_write_excel(self):
        api_table1 = ApiTable(
            id="t1",
            name="My Table",
            cells=[
                ApiCell(id="c1", row=1, column=1, cell_type=CellType.VALUE, value="Hello"),
                ApiCell(id="c2", row=1, column=2, cell_type=CellType.LINK, reference=[CellReference(table_id="t2", cell_address="A1")])
            ]
        )
        api_table2 = ApiTable(
            id="t2",
            name="Linked Table",
            cells=[ApiCell(id="c2_1", row=1, column=1, cell_type=CellType.VALUE, value="World")]
        )

        tables = [Table(api_table1), Table(api_table2)]
        output_file = "test_output.xlsx"

        writer = ExcelWriter(output_file)
        writer.write_tables(tables)

        self.assertTrue(os.path.exists(output_file))

        import openpyxl
        wb = openpyxl.load_workbook(output_file)
        self.assertIn("My Table", wb.sheetnames)
        self.assertIn("Linked Table", wb.sheetnames)

        ws1 = wb["My Table"]
        self.assertEqual(ws1["A1"].value, "Hello")
        self.assertEqual(ws1["B1"].value, "='Linked Table'!A1")

        wb.close()
        os.remove(output_file)

    def test_sanitize_sheet_name(self):
        writer = ExcelWriter("dummy")
        name = "Table:With/Invalid*Chars?"
        sanitized = writer._sanitize_sheet_name(name)
        self.assertEqual(sanitized, "TableWithInvalidChars")

if __name__ == '__main__':
    unittest.main()